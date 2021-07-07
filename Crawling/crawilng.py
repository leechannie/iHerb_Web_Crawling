import time
import openpyxl
import requests
import filecmp
import shutil
from openpyxl.styles import PatternFill, Color
from datetime import datetime
from bs4 import BeautifulSoup

# url을 담은 txt 파일 읽어오기 
def open_url():
    with open('C:/Users/Public/url.txt', 'rt', encoding='UTF8') as file:
        content = list()

        while True:
            sentence = file.readline().rstrip()

            if sentence:
                content.append(sentence)
            else:
                break

        return (content)

# 두 파일이 같은지 확인
def check_url(wb):
    flag = filecmp.cmp('C:/Users/Public/url.txt', 'C:/Users/Public/url_copy.txt')
    #print(flag)
    # 두 파일이 같지 않으면 url_copy에 url 파일 복사하기 
    if (flag == False):
        shutil.copyfile('C:/Users/이찬희/python2021/url.txt', 'C:/Users/이찬희/python2021/url_copy.txt' ) 
        add_exel(wb)
    else:
        pass

# BeautifulSoupd으로 특정 페이지를 연다.
def get_url(): 
    content = open_url()
    url_list = list()
    
    for i in  range (len(content)):
        try: 
            webpage = requests.get(content[i])
            soup = BeautifulSoup(webpage.content, "html.parser")
            url_list.append(soup)
        except:
            pass

    return (url_list)

#상품의 고유 번호를 가져온다.   
def get_name(): 
    url_list = get_url()
    name_list = list()
    for i in url_list: 
        name_list.append(i.find('span', itemprop="gtin12").text)
    
    return (name_list)

#상품의 재고 상황을 가져온다.
def scrapping():
    status_list = list()
    url_list = get_url()
     
    for i in url_list:
        try:
            status_list.append(i.find('div', 'text-danger stock-status-text').text.strip())

        except:
            status_list.append('재고 있음')
    return (status_list)

def make_exel(): 
    wb = openpyxl.Workbook()

    sheet1 = wb['Sheet']

    # 고유번호 값을 가져온다. 
    name_list = get_name()
    content = open_url()

    # 헤더를 추가한다. 
    sheet1.append(["날짜", "시간"])
    sheet1.column_dimensions['A'].width = 15

    # 고유번호로 상품의 이름을 보여주고, 하이퍼링크를 통해 누르면 상품 사이트와 연결되도록 한다. 
    for i in range (len(name_list)):
        sheet1.cell(row=1, column=i+3).value = '=HYPERLINK("{}", "{}")'.format(content[i], name_list[i])
        sheet1.cell(row=1, column=i+3).style = "Hyperlink" # 하이퍼링크처럼 파란색으로 보여주는 코드
        
    wb.save('재고확인.xlsx')

# url에 따라 다시 헤더 변경 
def add_exel(wb): 
    sheet1 = wb.active

    # 고유번호 값을 가져온다. 
    name_list = get_name()
    content = open_url()

    # 고유번호로 상품의 이름을 보여주고, 하이퍼링크를 통해 누르면 상품 사이트와 연결되도록 한다. 
    for i in range (len(name_list)):
        sheet1.cell(row=1, column=i+3).value = '=HYPERLINK("{}", "{}")'.format(content[i], name_list[i])
        sheet1.cell(row=1, column=i+3).style = "Hyperlink" # 하이퍼링크처럼 파란색으로 보여주는 코드
    wb.save('재고확인.xlsx')

def main(): 
    make_exel()
    wb = openpyxl.load_workbook('재고확인.xlsx')

    # 엑셀의 row를 위한 변수
    count = 1

    while True:  
        try:
            #시간 설정 
            now = datetime.now()
            date = "%s년 %s월 %s일" %(now.year, now.month, now.day)
            hour = "%s시 %s분" %(now.hour, now.minute)

            # 메모장에 url이 추가되는지 확인
            print("check url...")
            check_url(wb)
            
            # scrapping 함수를 호출 
            print("scrapping ... ")
            status_list = scrapping()

            sheet1 = wb.active
            sheet1.append([date, hour])
            # 각 셀에 상태 저장 
            for i in range (len(status_list)):
                if (status_list[i].find('품절') == 0):
                    #print(status_list[i])
                    sheet1.cell(row=count+1, column=i+3).value = status_list[i]
                    sheet1.cell(row=count+1, column=i+3).fill = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
                elif (status_list[i] == '재고 있음'):
                    sheet1.cell(row=count+1, column=i+3).value = status_list[i]
                    sheet1.cell(row=count+1, column=i+3).fill = PatternFill(start_color='0099CC00', end_color='0099CC00', fill_type='solid')

            print('wait...')
            #1분마다 정보를 갱신한다. 
            time.sleep(59)  
            try:
                count += 1 # 엑셀의 row를 위한 변수
                wb.save('재고확인.xlsx')
            except:
                pass
        except:
            print("Error")
            break

if __name__ == '__main__':
    main()